#!/usr/bin/env python3
"""
spreadsheet-of-doom.py — wide investigative tracking spreadsheet generator.

Walks every ./analysis/<domain>/findings.md plus ./analysis/correlation.md in
a case workspace, normalises each finding heading into one row, and emits:

    ./reports/spreadsheet-of-doom.csv   (always)
    ./reports/spreadsheet-of-doom.xlsx  (when openpyxl is importable)

Read-only against analysis artifacts. Stdlib-only for CSV; openpyxl is a soft
dependency for XLSX (try / except ImportError -> warn-and-continue).

USAGE
    python3 spreadsheet-of-doom.py <case-workspace-path>

The script must run from any CWD. The positional argument is the absolute or
relative path to the case workspace (the directory that contains analysis/,
exports/, reports/). All output paths are derived from that root.

CONTRACT (from DISCIPLINE.md, exec-briefing/SKILL.md, the issue body)
- Heading text is the Finding ID. Finding IDs that resolve back to a
  findings.md heading are required by the issue's acceptance criteria.
- Domain is the directory name under ./analysis/ that contains the file.
- Per-finding metadata is parsed from `**Field:** value` bullet lines under
  the heading, until the next heading or end of file.
- Description is the first sentence (period / newline) of the Finding field,
  or of the body if Finding is absent.
- Cells we cannot fill from disk are emitted blank — never invented.
- correlation.md contributes a Correlated findings cell for any heading whose
  body references one or more Finding IDs.

The XLSX form adds:
- frozen header row (A2)
- autofilter across the dimensions
- conditional formatting: Severity == high -> red, Status == open -> yellow

(See issue #2 for the full acceptance-criteria list.)
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Column schema. Order is the column order in CSV / XLSX. Keep stable;
# downstream consumers (analyst spreadsheets, QA row-count checks) rely on it.
# ---------------------------------------------------------------------------
COLUMNS = [
    "Finding ID",
    "Domain",
    "Evidence item",
    "Host",
    "User",
    "IP",
    "Event UTC",
    "Description",
    "MITRE technique",
    "Severity",
    "Confidence",
    "Status",
    "Correlated findings",
    "Source artifact path",
    "Hash / IOC",
    "Lead ID",
]

# Field-name -> column-name mapping. Keys are lowercased after stripping the
# bullet/asterisk markup. Multiple synonyms collapse to the same column so a
# finding written with "Artefact:" or "Artifact path:" still lands in
# "Source artifact path".
FIELD_TO_COLUMN = {
    "finding id": "Finding ID",
    "evidence": "Evidence item",
    "evidence item": "Evidence item",
    "evidence id": "Evidence item",
    "host": "Host",
    "hostname": "Host",
    "system": "Host",
    "user": "User",
    "username": "User",
    "account": "User",
    "ip": "IP",
    "ip address": "IP",
    "src ip": "IP",
    "source ip": "IP",
    "event utc": "Event UTC",
    "timestamp": "Event UTC",
    "time": "Event UTC",
    "utc": "Event UTC",
    "when": "Event UTC",
    "finding": "Description",
    "description": "Description",
    "summary": "Description",
    "mitre": "MITRE technique",
    "mitre technique": "MITRE technique",
    "mitre attack": "MITRE technique",
    "att&ck": "MITRE technique",
    "technique": "MITRE technique",
    "severity": "Severity",
    "priority": "Severity",
    "confidence": "Confidence",
    "status": "Status",
    "correlated": "Correlated findings",
    "correlated findings": "Correlated findings",
    "related": "Correlated findings",
    "artifact": "Source artifact path",
    "artefact": "Source artifact path",
    "artifact path": "Source artifact path",
    "source": "Source artifact path",
    "source artifact": "Source artifact path",
    "source artifact path": "Source artifact path",
    "path": "Source artifact path",
    "file": "Source artifact path",
    "hash": "Hash / IOC",
    "hash/ioc": "Hash / IOC",
    "hash / ioc": "Hash / IOC",
    "sha256": "Hash / IOC",
    "md5": "Hash / IOC",
    "ioc": "Hash / IOC",
    "indicator": "Hash / IOC",
    "lead": "Lead ID",
    "lead id": "Lead ID",
}

# Pre-compiled regex: a level-2 markdown heading (## ...) is one finding.
HEADING_RE = re.compile(r"^##\s+(.+?)\s*$")
# Field bullets: "- **Field:** value" or "* **Field:** value", with or without
# trailing emphasis and arbitrary whitespace.
FIELD_BULLET_RE = re.compile(
    r"^\s*[-*]\s+\*\*([^*]+?)\*\*\s*[:\-]?\s*(.*)$"
)
# Catch a Finding-ID line written without bold (e.g. "Finding ID: L-EV01-...").
PLAIN_FIELD_RE = re.compile(r"^\s*([A-Za-z][A-Za-z &/]+?)\s*:\s*(.+?)\s*$")
# Lead-ID pattern as defined by ORCHESTRATE.md (L-EV01-memory-01, L-CORR-04, etc.).
LEAD_ID_RE = re.compile(r"\bL-(?:EV\d+-[a-z0-9-]+|CORR-\d+|BASELINE-[a-z0-9-]+)-?\d*\b")
# Finding-ID heuristic: anything inside `backticks` or matching the Lead-ID
# shape is treated as a Finding ID reference when scanning correlation rows.
BACKTICK_ID_RE = re.compile(r"`([^`]+)`")


def first_sentence(text: str) -> str:
    """First sentence of the description body.

    A sentence terminator is `.`, `!`, or `?` followed by whitespace AND a
    capital letter / digit / end-of-string. Plain `.<digit>` (e.g. inside an
    IP address `198.51.100.7` or a version `1.2.3`) does NOT terminate a
    sentence -- this matters because forensic finding text is dense with
    dotted-decimal IPs and hashes. Newlines hard-terminate.
    """
    if not text:
        return ""
    cleaned = text.strip().splitlines()[0].strip()
    # Walk forward looking for a real sentence boundary.
    n = len(cleaned)
    for i, ch in enumerate(cleaned):
        if ch not in ".!?":
            continue
        nxt = cleaned[i + 1] if i + 1 < n else ""
        nxt2 = cleaned[i + 2] if i + 2 < n else ""
        if not nxt:
            return cleaned[: i + 1].strip()
        if nxt.isspace() and (not nxt2 or nxt2.isupper() or nxt2.isdigit() or nxt2 in "(\"'"):
            return cleaned[: i + 1].strip()
    return cleaned


def normalise_field(name: str) -> str | None:
    """Map a parsed field-name token to a canonical column name."""
    if not name:
        return None
    key = name.strip().lower().rstrip(":").strip()
    return FIELD_TO_COLUMN.get(key)


def parse_findings_file(path: Path, domain: str) -> list[dict]:
    """Walk one findings.md, return a row dict per heading.

    Heading text becomes Finding ID (unless an explicit ``Finding ID:`` field
    overrides it). The body's bullet fields populate the rest of the row.
    """
    rows: list[dict] = []
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError as e:
        sys.stderr.write(f"warning: cannot read {path}: {e}\n")
        return rows

    current: dict | None = None
    body_lines: list[str] = []

    def flush() -> None:
        if current is None:
            return
        # Description is built last from accumulated body if not explicitly set.
        if not current.get("Description"):
            # First non-bullet line of the body, or the Finding bullet's text.
            for line in body_lines:
                stripped = line.strip()
                if not stripped or stripped.startswith(("-", "*", "#", ">")):
                    continue
                current["Description"] = first_sentence(stripped)
                break
        else:
            current["Description"] = first_sentence(current["Description"])
        # Capture lead IDs anywhere in the body if no explicit Lead ID column.
        if not current.get("Lead ID"):
            joined = "\n".join(body_lines)
            m = LEAD_ID_RE.search(joined)
            if m:
                current["Lead ID"] = m.group(0)
        rows.append(current)

    for raw in text.splitlines():
        m = HEADING_RE.match(raw)
        if m:
            # New heading -> flush the previous finding, start fresh.
            flush()
            heading = m.group(1).strip()
            current = {col: "" for col in COLUMNS}
            current["Finding ID"] = heading
            current["Domain"] = domain
            body_lines = []
            continue

        if current is None:
            # Pre-amble before the first heading -- ignore.
            continue

        body_lines.append(raw)
        # Try the bolded-bullet pattern first.
        bm = FIELD_BULLET_RE.match(raw)
        if bm:
            field, value = bm.group(1), bm.group(2)
            col = normalise_field(field)
            if col and not current.get(col):
                current[col] = value.strip()
            continue
        # Fall back to "Field: value" plain lines (e.g. analyst-tags rows).
        pm = PLAIN_FIELD_RE.match(raw)
        if pm:
            field, value = pm.group(1), pm.group(2)
            col = normalise_field(field)
            if col and not current.get(col):
                current[col] = value.strip()

    flush()
    return rows


def parse_correlation(path: Path, finding_ids: set[str]) -> dict[str, list[str]]:
    """Walk correlation.md, return {finding_id: [related_id, ...]}.

    A correlation entry is any markdown heading or list item whose body
    references two or more known finding IDs. Each referenced ID gets the
    other referenced IDs added to its Correlated-findings cell.
    """
    pairs: dict[str, set[str]] = {}
    if not path.exists():
        return {}
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return {}

    # Split into blocks at headings; each block is one correlation entry.
    blocks: list[str] = []
    buf: list[str] = []
    for line in text.splitlines():
        if HEADING_RE.match(line):
            if buf:
                blocks.append("\n".join(buf))
                buf = []
        buf.append(line)
    if buf:
        blocks.append("\n".join(buf))

    # Also break into pipe-table rows so a correlation matrix table contributes.
    extra: list[str] = []
    for line in text.splitlines():
        if line.strip().startswith("|") and line.count("|") >= 2:
            extra.append(line)
    blocks.extend(extra)

    for block in blocks:
        # Collect every Finding-ID-shaped token in this block: backticked refs
        # and plain Lead-ID-pattern refs. Restrict to IDs we have on disk so
        # we never write a phantom pointer into the spreadsheet.
        candidates: set[str] = set()
        for m in BACKTICK_ID_RE.finditer(block):
            tok = m.group(1).strip()
            if tok in finding_ids:
                candidates.add(tok)
        for m in LEAD_ID_RE.finditer(block):
            tok = m.group(0)
            if tok in finding_ids:
                candidates.add(tok)
        # Heading-text refs: if the block heading is itself a known Finding ID,
        # treat every other ID in the block as related to it.
        head = HEADING_RE.match(block.split("\n", 1)[0]) if block else None
        if head and head.group(1).strip() in finding_ids:
            candidates.add(head.group(1).strip())

        if len(candidates) < 2:
            continue
        for fid in candidates:
            pairs.setdefault(fid, set()).update(c for c in candidates if c != fid)

    return {k: sorted(v) for k, v in pairs.items()}


def collect_rows(case_root: Path) -> list[dict]:
    """Walk all analysis/<domain>/findings.md and merge correlation.md refs."""
    analysis_dir = case_root / "analysis"
    if not analysis_dir.is_dir():
        sys.stderr.write(f"error: {analysis_dir} does not exist\n")
        sys.exit(2)

    rows: list[dict] = []
    # Sorted glob keeps row order deterministic across runs.
    for findings_path in sorted(analysis_dir.glob("*/findings.md")):
        domain = findings_path.parent.name
        rows.extend(parse_findings_file(findings_path, domain))

    finding_ids = {r["Finding ID"] for r in rows if r.get("Finding ID")}
    correlations = parse_correlation(analysis_dir / "correlation.md", finding_ids)
    for r in rows:
        related = correlations.get(r["Finding ID"], [])
        if related and not r.get("Correlated findings"):
            r["Correlated findings"] = ", ".join(related)
    return rows


def write_csv(rows: list[dict], out_path: Path) -> None:
    """Always-writable CSV output. UTF-8, newline='' for cross-platform."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in COLUMNS})


def write_xlsx(rows: list[dict], out_path: Path) -> bool:
    """Try-best XLSX output. Returns False if openpyxl is unavailable."""
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.stderr.write(
            "warning: openpyxl not installed; skipping spreadsheet-of-doom.xlsx "
            "(CSV still produced)\n"
        )
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Spreadsheet of Doom"
    ws.append(COLUMNS)
    # Header style: bold so the column titles read at a glance.
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])

    # Frozen header row (A2 means 'rows above row 2 are frozen').
    ws.freeze_panes = "A2"
    # Autofilter spans the populated dimensions (header + data).
    ws.auto_filter.ref = ws.dimensions

    # Conditional formatting -- minimal: high severity = red, open status = yellow.
    if rows:
        last_row = len(rows) + 1
        sev_idx = COLUMNS.index("Severity") + 1
        sta_idx = COLUMNS.index("Status") + 1
        sev_letter = get_column_letter(sev_idx)
        sta_letter = get_column_letter(sta_idx)
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        # Case-insensitive match via FormulaRule (CellIsRule is exact-string).
        ws.conditional_formatting.add(
            f"{sev_letter}2:{sev_letter}{last_row}",
            FormulaRule(
                formula=[f'EXACT(UPPER({sev_letter}2),"HIGH")'],
                fill=red,
            ),
        )
        ws.conditional_formatting.add(
            f"{sta_letter}2:{sta_letter}{last_row}",
            FormulaRule(
                formula=[f'EXACT(LOWER({sta_letter}2),"open")'],
                fill=yellow,
            ),
        )
        # Defensive auto-width: bound to a reasonable max so a long Description
        # cell does not blow up the column to 1000+ chars wide.
        for idx, col in enumerate(COLUMNS, start=1):
            longest = max(
                [len(col)] + [len(str(r.get(col, ""))) for r in rows]
            )
            ws.column_dimensions[get_column_letter(idx)].width = min(60, max(12, longest + 2))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return True


def resolve_case_root(arg: str) -> Path:
    """Expand the positional arg to an absolute path and verify it's a case dir."""
    root = Path(arg).expanduser().resolve()
    if not root.is_dir():
        sys.stderr.write(f"error: {root} is not a directory\n")
        sys.exit(2)
    if not (root / "analysis").is_dir():
        sys.stderr.write(
            f"error: {root} does not look like a case workspace "
            "(no ./analysis/ directory)\n"
        )
        sys.exit(2)
    return root


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Generate the Spreadsheet of Doom (CSV always, XLSX when openpyxl "
            "is available) from a case workspace's analysis artifacts."
        )
    )
    parser.add_argument(
        "case_workspace",
        help="Path to ./cases/<CASE_ID>/ (the directory containing analysis/).",
    )
    args = parser.parse_args(argv)

    case_root = resolve_case_root(args.case_workspace)
    rows = collect_rows(case_root)
    csv_path = case_root / "reports" / "spreadsheet-of-doom.csv"
    xlsx_path = case_root / "reports" / "spreadsheet-of-doom.xlsx"

    write_csv(rows, csv_path)
    xlsx_written = write_xlsx(rows, xlsx_path)

    print(f"Spreadsheet of Doom: {len(rows)} row(s)")
    print(f"  CSV : {csv_path}")
    if xlsx_written:
        print(f"  XLSX: {xlsx_path}")
    else:
        print("  XLSX: (skipped -- openpyxl not installed)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
